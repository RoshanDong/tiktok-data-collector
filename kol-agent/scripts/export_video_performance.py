#!/usr/bin/env python3
"""
TikTok Video Performance Data Exporter
批量导出近7天每天的全量视频数据到 Excel 表格
"""

import json
import time
import hmac
import hashlib
import logging
import os
import argparse
from datetime import datetime, timedelta
from typing import Optional, Dict, Any, List
from functools import wraps
from collections import defaultdict

import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# 钉钉上传模块
from dingtalk_uploader import upload_excel

# Configuration path
CONFIG_PATH = ".kol-agent/tiktok-config.json"
DEFAULT_OUTPUT_DIR = ".kol-agent/tiktok-data/exports"
PAGE_SIZE = 100  # API 最大 page_size
DEFAULT_UPLOAD = True  # 默认上传到钉钉

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(levelname)s - %(message)s',
    datefmt='%H:%M:%S'
)
logger = logging.getLogger(__name__)


# ============================================================================
# 依赖检查
# ============================================================================
def check_dependencies():
    """检查依赖是否已安装"""
    try:
        import requests
        import openpyxl
        import tenacity
        logger.info("✓ 依赖检查通过: requests, openpyxl, tenacity")
        return True
    except ImportError as e:
        logger.error(f"✗ 依赖缺失: {e}")
        logger.error("请运行: pip install requests openpyxl tenacity")
        return False


# ============================================================================
# Token 刷新函数 (Phase 2)
# ============================================================================
def refresh_token(config: dict) -> dict:
    """
    使用 refresh_token 刷新 access_token
    """
    api_config = config.get("tiktok_api", {})

    params = {
        "app_key": api_config.get("client_key"),
        "app_secret": api_config.get("client_secret"),
        "refresh_token": api_config.get("refresh_token"),
        "grant_type": "refresh_token"
    }

    url = "https://auth.tiktok-shops.com/api/v2/token/refresh?" + "&".join(
        f"{k}={v}" for k, v in params.items()
    )

    logger.info("刷新 access_token...")

    try:
        resp = requests.get(url, timeout=30)
        result = resp.json()

        if result.get("code") == 0:
            data = result.get("data", {})
            api_config["access_token"] = data.get("access_token")
            api_config["refresh_token"] = data.get("refresh_token")
            api_config["token_expire_in"] = data.get("access_token_expire_in")

            # Save updated config
            with open(CONFIG_PATH, "w") as f:
                json.dump(config, f, indent=2)

            logger.info("✓ Token 刷新成功")
            return config
        else:
            logger.error(f"✗ Token 刷新失败: {result.get('message')}")
            return None

    except Exception as e:
        logger.error(f"✗ Token 刷新异常: {e}")
        return None


# ============================================================================
# 指数退避重试装饰器 (Phase 2)
# ============================================================================
def retry_with_exponential_backoff(max_attempts: int = 3, initial_delay: float = 1.0, max_delay: float = 30.0):
    """
    指数退避重试装饰器
    - 初始延迟 1 秒，每次重试翻倍，最高 30 秒
    - 最多 3 次尝试
    """
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            delay = initial_delay
            last_exception = None

            for attempt in range(1, max_attempts + 1):
                try:
                    result = func(*args, **kwargs)
                    if attempt > 1:
                        logger.info(f"✓ 重试成功 (第 {attempt} 次尝试)")
                    return result

                except Exception as e:
                    last_exception = e
                    logger.warning(f"✗ 调用失败 (第 {attempt}/{max_attempts} 次): {e}")

                    if attempt < max_attempts:
                        logger.info(f"  等待 {delay:.1f} 秒后重试...")
                        time.sleep(delay)
                        delay = min(delay * 2, max_delay)
                    else:
                        logger.error(f"✗ 超过最大重试次数 ({max_attempts})")

            raise last_exception

        return wrapper
    return decorator


# ============================================================================
# 响应验证函数 (Phase 2)
# ============================================================================
class APIError(Exception):
    """API 调用错误异常"""
    def __init__(self, code: int, message: str, request_id: str = None):
        self.code = code
        self.message = message
        self.request_id = request_id
        super().__init__(f"API Error {code}: {message}")


def validate_response(response: dict, required_fields: List[str] = None) -> bool:
    """
    验证 API 响应
    - 检查 code == 0
    - 检查 message == "success"
    - 可选：检查必需字段存在
    """
    code = response.get("code")
    message = response.get("message")
    request_id = response.get("request_id")

    if code != 0:
        raise APIError(code, message, request_id)

    if message and message.lower() != "success":
        logger.warning(f"响应消息异常: {message}")

    # Check required fields (Phase 2 - missing field handling)
    if required_fields:
        data = response.get("data", {})
        for field in required_fields:
            if field not in data:
                raise APIError(0, f"缺少必需字段: {field}", request_id)

    return True


# ============================================================================
# API 客户端 (Phase 2 + Phase 3 - pagination)
# ============================================================================
class VideoPerformanceAPIClient:
    """TikTok Shop Video Performance API 客户端"""

    BASE_URL = "https://open-api.tiktokglobalshop.com"

    def __init__(self, config_path: str = CONFIG_PATH):
        with open(config_path, "r") as f:
            self.config = json.load(f)

        api_config = self.config.get("tiktok_api", {})
        self.app_key = api_config.get("client_key")
        self.app_secret = api_config.get("client_secret")
        self.access_token = api_config.get("access_token")
        self.shop_cipher = api_config.get("shop_cipher")

    def _make_request(self, method: str, path: str, query_params: dict = None) -> dict:
        """基础 API 请求方法"""
        timestamp = int(time.time())

        # Build query params
        params = query_params.copy() if query_params else {}
        params["app_key"] = self.app_key
        params["timestamp"] = timestamp

        # Generate signature
        sign_str = self.app_secret + path
        for k in sorted(params.keys()):
            sign_str += k + str(params[k])
        sign_str += self.app_secret

        sign = hmac.new(self.app_secret.encode(), sign_str.encode(), hashlib.sha256).hexdigest()

        # Build URL
        url = f"{self.BASE_URL}{path}?app_key={self.app_key}&"
        if query_params:
            for k, v in query_params.items():
                url += f"{k}={v}&"
        url += f"timestamp={timestamp}&sign={sign}"

        headers = {
            "Content-Type": "application/json",
            "x-tts-access-token": self.access_token
        }

        start_time = time.time()
        resp = requests.get(url, headers=headers)
        elapsed = time.time() - start_time

        logger.debug(f"→ GET {url} ({elapsed:.2f}s)")

        result = resp.json()
        return result

    @retry_with_exponential_backoff()
    def call_get_shop_video_performance(self, start_date_ge: str, end_date_lt: str,
                                        page_token: str = None,
                                        page_size: int = PAGE_SIZE) -> dict:
        """
        Get Shop Video Performance List (版本 202605)
        支持分页获取
        """
        path = "/analytics/202605/shop_videos/performance"

        query_params = {
            "shop_cipher": self.shop_cipher,
            "start_date_ge": start_date_ge,
            "end_date_lt": end_date_lt,
            "sort_field": "gmv",
            "currency": "USD",
            "page_size": page_size,
            "sort_order": "DESC",
            "account_type": "ALL",
        }

        if page_token:
            query_params["page_token"] = page_token

        result = self._make_request("GET", path, query_params)
        validate_response(result, required_fields=["videos", "total_count"])
        return result

    def fetch_all_videos_for_day(self, start_date_ge: str, end_date_lt: str) -> tuple:
        """
        获取指定日期范围内的所有视频 (Phase 3 - pagination loop)
        返回: (videos_list, total_count)
        """
        all_videos = []
        page_token = None
        total_count = None

        logger.info(f"  开始获取 {start_date_ge} 的视频数据 (page_size={PAGE_SIZE})")

        while True:
            result = self.call_get_shop_video_performance(
                start_date_ge=start_date_ge,
                end_date_lt=end_date_lt,
                page_token=page_token,
                page_size=PAGE_SIZE
            )

            data = result.get("data", {})
            videos = data.get("videos", [])
            total_count = data.get("total_count", 0)
            next_token = data.get("next_page_token")

            all_videos.extend(videos)
            logger.info(f"  已获取 {len(all_videos)}/{total_count} 条视频")

            if not next_token or len(all_videos) >= total_count:
                break

            page_token = next_token
            time.sleep(0.5)  # 避免请求过快

        return all_videos, total_count

    def get_latest_available_date(self) -> str:
        """
        获取 API 最新可用日期
        通过调用一次 API 获取 latest_available_date
        """
        # 使用一个合理的日期范围来触发 latest_available_date 返回
        today = datetime.now()
        start_date = (today - timedelta(days=30)).strftime("%Y-%m-%d")
        end_date = (today + timedelta(days=2)).strftime("%Y-%m-%d")

        result = self._make_request(
            "GET",
            "/analytics/202605/shop_videos/performance",
            {
                "shop_cipher": self.shop_cipher,
                "start_date_ge": start_date,
                "end_date_lt": end_date,
                "sort_field": "gmv",
                "currency": "USD",
                "page_size": 1,
                "sort_order": "DESC",
                "account_type": "ALL",
            }
        )
        validate_response(result, required_fields=["latest_available_date"])
        latest_date = result.get("data", {}).get("latest_available_date")
        logger.info(f"  API 最新可用日期: {latest_date}")
        return latest_date


# ============================================================================
# Excel 写入器 (Phase 3)
# ============================================================================
class ExcelWriter:
    """Excel 写入器 - 按照 data-model.md schema 输出"""

    COLUMNS = [
        ("A", "id", "Video ID"),
        ("B", "title", "Title"),
        ("C", "username", "Username"),
        ("D", "creator_user_name", "Creator User Name"),
        ("E", "creator_nick_name", "Creator Nick Name"),
        ("F", "creator_author_type", "Creator Author Type"),
        ("G", "video_post_time", "Video Post Time"),
        ("H", "duration", "Duration (sec)"),
        ("I", "hash_tags", "Hashtags"),
        ("J", "gmv_amount", "GMV Amount"),
        ("K", "gmv_currency", "GMV Currency"),
        ("L", "gpm_amount", "GPM Amount"),
        ("M", "gpm_currency", "GPM Currency"),
        ("N", "avg_customers", "Avg Customers"),
        ("O", "sku_orders", "SKU Orders"),
        ("P", "items_sold", "Items Sold"),
        ("Q", "views", "Views"),
        ("R", "click_through_rate", "Click Through Rate"),
        ("S", "products", "Products"),
    ]

    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Videos"

    def write_header(self):
        """写入表头"""
        for col_letter, field_name, header_text in self.COLUMNS:
            cell = self.worksheet[f"{col_letter}1"]
            cell.value = header_text
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGN

        # 设置列宽
        widths = {
            "A": 20, "B": 60, "C": 20, "D": 20, "E": 25,
            "F": 20, "G": 20, "H": 12, "I": 40, "J": 12,
            "K": 12, "L": 12, "M": 12, "N": 15, "O": 12,
            "P": 12, "Q": 12, "R": 20, "S": 60
        }
        for col, width in widths.items():
            self.worksheet.column_dimensions[col].width = width

    def _get_nested_field(self, video: dict, field_path: str, default=None):
        """获取嵌套字段 (如 creator.user_name)"""
        parts = field_path.split(".")
        value = video
        for part in parts:
            if isinstance(value, dict):
                value = value.get(part, default)
            else:
                return default
        return value if value is not None else default

    def _format_value(self, video: dict, field_name: str) -> Any:
        """根据字段名格式化值"""
        if field_name == "hash_tags":
            tags = video.get("hash_tags", [])
            return ", ".join(tags) if tags else ""
        elif field_name == "products":
            prods = video.get("products", [])
            return "; ".join([f"{p.get('id', '')}: {p.get('name', '')}" for p in prods]) if prods else ""
        elif field_name in ("gmv_amount", "gpm_amount"):
            obj = video.get(field_name.split("_")[0], {})
            return obj.get("amount", "0") if obj else "0"
        elif field_name in ("gmv_currency", "gpm_currency"):
            obj = video.get(field_name.split("_")[0], {})
            return obj.get("currency", "USD") if obj else "USD"
        elif field_name == "creator_user_name":
            return self._get_nested_field(video, "creator.user_name", "")
        elif field_name == "creator_nick_name":
            return self._get_nested_field(video, "creator.nick_name", "")
        elif field_name == "creator_author_type":
            return self._get_nested_field(video, "creator.author_type", "")
        elif field_name == "video_post_time":
            return video.get("video_post_time", "")
        elif field_name == "duration":
            return video.get("duration", 0)
        elif field_name == "click_through_rate":
            return video.get("click_through_rate", "")
        else:
            return video.get(field_name, video.get("id", ""))

    def write_video(self, video: dict, row: int):
        """写入单条视频数据"""
        for col_letter, field_name, _ in self.COLUMNS:
            value = self._format_value(video, field_name)
            self.worksheet[f"{col_letter}{row}"] = value

    def save(self):
        """保存文件"""
        self.workbook.save(self.file_path)
        logger.info(f"  Excel 文件已保存: {self.file_path}")


# ============================================================================
# 主导出函数 (Phase 3 + Phase 5)
# ============================================================================
def export_date(client: VideoPerformanceAPIClient, date: str, output_dir: str) -> dict:
    """
    导出指定日期的视频数据
    返回统计信息
    """
    # 计算日期范围 (该日期整天)
    start_date = date
    next_day = (datetime.strptime(date, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")

    logger.info(f"  正在导出: {start_date}")

    try:
        # 获取所有视频
        videos, total_count = client.fetch_all_videos_for_day(start_date, next_day)

        if not videos:
            logger.warning(f"  {start_date} 无视频数据，跳过")
            return {"date": date, "status": "skipped", "count": 0}

        # 写入 Excel
        file_name = f"video_performance_{date}.xlsx"
        file_path = os.path.join(output_dir, file_name)

        writer = ExcelWriter(file_path)
        writer.write_header()

        for i, video in enumerate(videos, start=2):
            writer.write_video(video, i)

        writer.save()

        logger.info(f"  ✓ {start_date} 导出成功: {len(videos)} 条视频 → {file_name}")

        return {
            "date": date,
            "status": "success",
            "count": len(videos),
            "file": file_path
        }

    except APIError as e:
        logger.error(f"  ✗ {start_date} API 错误: {e}")
        return {"date": date, "status": "error", "error": str(e)}
    except Exception as e:
        logger.error(f"  ✗ {start_date} 导出异常: {e}")
        return {"date": date, "status": "error", "error": str(e)}


def get_date_range(days: int = 7) -> list:
    """
    计算日期范围 (Phase 3)
    返回过去 N 天的日期列表
    """
    today = datetime.now()
    dates = []
    for i in range(1, days + 1):
        date = (today - timedelta(days=i)).strftime("%Y-%m-%d")
        dates.append(date)
    return dates


def is_latest_date_exported(output_dir: str, date: str) -> bool:
    """检查最新一天的 Excel 文件是否已存在"""
    file_name = f"video_performance_{date}.xlsx"
    file_path = os.path.join(output_dir, file_name)
    return os.path.exists(file_path)


def main():
    """主函数"""
    logger.info("=" * 60)
    logger.info("TikTok Video Performance Data Exporter")
    logger.info("=" * 60)

    # 依赖检查
    if not check_dependencies():
        return

    # CLI 参数解析 (Phase 4)
    parser = argparse.ArgumentParser(description="导出 TikTok 视频表现数据")
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR,
                        help=f"输出目录 (默认: {DEFAULT_OUTPUT_DIR})")
    parser.add_argument("--days", type=int, default=None,
                        help="导出天数 (默认: 1，仅抓取最新可用日期)")
    parser.add_argument("--start-date",
                        help="开始日期 (YYYY-MM-DD)，覆盖默认行为")
    parser.add_argument("--end-date",
                        help="结束日期 (YYYY-MM-DD)，覆盖默认行为")
    parser.add_argument("--upload", action="store_true", default=DEFAULT_UPLOAD,
                        help="导出后上传到钉钉 (默认开启)")
    parser.add_argument("--no-upload", action="store_true",
                        help="禁用上传到钉钉")
    parser.add_argument("--upload-only",
                        help="仅上传指定日期的 Excel 文件 (YYYY-MM-DD)")
    args = parser.parse_args()

    # 验证输出目录 (Phase 4)
    output_dir = args.output_dir
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        logger.info(f"✓ 创建输出目录: {output_dir}")
    elif not os.access(output_dir, os.W_OK):
        logger.error(f"✗ 输出目录不可写: {output_dir}")
        return

    # 初始化 API 客户端
    client = VideoPerformanceAPIClient()

    # 获取 API 最新可用日期
    logger.info("获取 API 最新可用日期...")
    latest_date = client.get_latest_available_date()

    # 计算日期范围
    if args.start_date and args.end_date:
        # 自定义日期范围
        start_dt = datetime.strptime(args.start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(args.end_date, "%Y-%m-%d")
        dates = []
        current = start_dt
        while current <= end_dt:
            dates.append(current.strftime("%Y-%m-%d"))
            current += timedelta(days=1)
        logger.info(f"将导出 {len(dates)} 天的数据: {dates[0]} ~ {dates[-1]}")
    elif args.days:
        # 自定义天数
        dates = get_date_range(args.days)
        logger.info(f"将导出 {len(dates)} 天的数据: {dates[0]} ~ {dates[-1]}")
    else:
        # 默认行为：只抓取最新可用日期一天
        dates = [latest_date]
        logger.info(f"将导出最新一天的数据: {latest_date}")

    # 主导出循环 (Phase 3 + Phase 5)
    results = []
    total_records = 0
    success_count = 0
    error_count = 0
    skipped_count = 0

    for date in dates:
        # 检查是否已存在（仅针对最新日期）
        if date == latest_date and is_latest_date_exported(output_dir, date):
            logger.info(f"  ⏭ {date} 数据已存在，跳过抓取")
            results.append({"date": date, "status": "skipped", "count": 0})
            skipped_count += 1
            continue

        result = export_date(client, date, output_dir)
        results.append(result)

        if result["status"] == "success":
            total_records += result["count"]
            success_count += 1
        elif result["status"] == "error":
            error_count += 1

    # 最终统计 (Phase 5)
    logger.info("\n" + "=" * 60)
    logger.info("导出完成")
    logger.info("=" * 60)
    logger.info(f"  成功: {success_count} 天")
    logger.info(f"  失败: {error_count} 天")
    logger.info(f"  跳过: {skipped_count} 天")
    logger.info(f"  总记录数: {total_records}")
    logger.info(f"  输出目录: {output_dir}")

    # 处理仅上传模式
    if args.upload_only:
        upload_date = args.upload_only
        file_name = f"video_performance_{upload_date}.xlsx"
        file_path = os.path.join(output_dir, file_name)

        if not os.path.exists(file_path):
            logger.error(f"✗ 文件不存在: {file_path}")
        else:
            logger.info(f"\n{'=' * 60}")
            logger.info(f"上传 {upload_date} 数据到钉钉...")
            upload_result = upload_excel(file_path, upload_date)
            if upload_result.get("success"):
                logger.info(f"✓ 上传成功: media_id = {upload_result.get('media_id')}")
            else:
                logger.error(f"✗ 上传失败: {upload_result.get('error')}")
        return

    # 上传到钉钉 (如果有成功导出的文件)
    enable_upload = args.upload and not args.no_upload
    if enable_upload and success_count > 0:
        # 上传最新一天的数据
        latest_success = None
        for r in results:
            if r["status"] == "success":
                latest_success = r

        if latest_success:
            logger.info(f"\n{'=' * 60}")
            logger.info(f"上传 {latest_success['date']} 数据到钉钉...")
            upload_result = upload_excel(latest_success["file"], latest_success["date"])
            if upload_result.get("success"):
                logger.info(f"✓ 上传成功: media_id = {upload_result.get('media_id')}")
            else:
                logger.error(f"✗ 上传失败: {upload_result.get('error')}")

    if error_count > 0:
        logger.warning("失败日期:")
        for r in results:
            if r["status"] == "error":
                logger.warning(f"  {r['date']}: {r.get('error', 'Unknown')}")


if __name__ == "__main__":
    main()