"""
Excel Writer Module for TikTok Shop Data Fetch Automation
Handles Excel file creation and monthly data accumulation.
"""

import json
from datetime import datetime
from pathlib import Path
from typing import Optional, List
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from models import DailySalesData


class ExcelWriter:
    """Handles Excel file creation and monthly data accumulation"""

    def __init__(self, data_dir: Path):
        """
        Initialize Excel writer with data directory.

        Args:
            data_dir: Directory path for Excel files (e.g., .kol-agent/tiktok-data/)
        """
        self.data_dir = Path(data_dir)
        self._ensure_data_dir()

    def _ensure_data_dir(self):
        """Ensure data directory exists"""
        self.data_dir.mkdir(parents=True, exist_ok=True)

    def _get_filename(self, year_month: str) -> Path:
        """
        Get Excel filename for a given year-month.

        Args:
            year_month: Year-month string in YYYY-MM format

        Returns:
            Path to the Excel file
        """
        return self.data_dir / f"tiktok-sales-{year_month}.xlsx"

    def _get_sheet_name(self, year_month: str) -> str:
        """
        Get Excel sheet name for a given year-month.

        Args:
            year_month: Year-month string in YYYY-MM format

        Returns:
            Sheet name (e.g., "2026-06")
        """
        return year_month  # YYYY-MM format works as sheet name

    def _create_headers(self, worksheet):
        """Create column headers with formatting"""
        headers = ["Date", "Order Count", "Total Revenue", "Top Products"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    def _get_or_create_workbook(self, filename: Path, year_month: str) -> Workbook:
        """
        Get existing workbook or create new one with formatted headers.

        Args:
            filename: Path to Excel file
            year_month: Year-month for sheet naming

        Returns:
            Workbook object
        """
        sheet_name = self._get_sheet_name(year_month)

        if filename.exists():
            # Load existing workbook
            wb = load_workbook(filename)
            # Create sheet if it doesn't exist
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(title=sheet_name)
                self._create_headers(ws)
            else:
                ws = wb[sheet_name]
        else:
            # Create new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            self._create_headers(ws)

        return wb

    def write_daily_data(self, sales_data: DailySalesData) -> Path:
        """
        Write daily sales data to Excel file.

        Args:
            sales_data: DailySalesData object to write

        Returns:
            Path to the Excel file that was written
        """
        # Extract year-month from date
        year_month = sales_data.date[:7]  # YYYY-MM
        filename = self._get_filename(year_month)

        # Get or create workbook
        wb = self._get_or_create_workbook(filename, year_month)
        ws = wb[self._get_sheet_name(year_month)]

        # Find next empty row
        next_row = ws.max_row + 1
        if ws.max_row == 1:  # Only header row exists
            next_row = 2

        # Write data row
        row_data = sales_data.to_excel_row()
        ws.cell(row=next_row, column=1).value = row_data["Date"]
        ws.cell(row=next_row, column=2).value = row_data["Order Count"]
        ws.cell(row=next_row, column=3).value = row_data["Total Revenue"]
        ws.cell(row=next_row, column=4).value = row_data["Top Products"]
        ws.cell(row=next_row, column=4).alignment = Alignment(wrap_text=True)

        # Auto-adjust column widths
        for col in range(1, 5):
            ws.column_dimensions[chr(64 + col)].width = 20

        # Save workbook
        wb.save(filename)
        return filename

    def read_monthly_data(self, year_month: str) -> List[DailySalesData]:
        """
        Read all daily sales data for a given month.

        Args:
            year_month: Year-month string in YYYY-MM format

        Returns:
            List of DailySalesData objects
        """
        filename = self._get_filename(year_month)
        if not filename.exists():
            return []

        wb = load_workbook(filename, data_only=True)
        ws = wb[self._get_sheet_name(year_month)]

        sales_list = []
        for row in range(2, ws.max_row + 1):  # Skip header
            date = ws.cell(row=row, column=1).value
            order_count = ws.cell(row=row, column=2).value
            total_revenue = ws.cell(row=row, column=3).value
            top_products_json = ws.cell(row=row, column=4).value

            if date:
                # Parse top products JSON
                top_products = []
                if top_products_json:
                    try:
                        products_data = json.loads(top_products_json)
                        from models import ProductSummary
                        top_products = [ProductSummary.from_dict(p) for p in products_data]
                    except json.JSONDecodeError:
                        pass

                sales_list.append(DailySalesData(
                    date=str(date),
                    orderCount=int(order_count or 0),
                    totalRevenue=float(total_revenue or 0.0),
                    topProducts=top_products
                ))

        return sales_list

    def file_exists(self, year_month: str) -> bool:
        """Check if monthly Excel file exists"""
        return self._get_filename(year_month).exists()


def get_excel_writer(config: dict) -> ExcelWriter:
    """
    Factory function to create an ExcelWriter from config dict.

    Args:
        config: Configuration dictionary with 'storage.data_dir' path

    Returns:
        ExcelWriter instance
    """
    data_dir = config.get('storage', {}).get('data_dir', '.kol-agent/tiktok-data')
    return ExcelWriter(Path(data_dir))


if __name__ == "__main__":
    # Test Excel writer
    from models import DailySalesData, ProductSummary

    test_data_dir = Path(".kol-agent/tiktok-data")
    writer = ExcelWriter(test_data_dir)

    # Test data
    sales_data = DailySalesData(
        date="2026-06-09",
        orderCount=156,
        totalRevenue=4523.50,
        topProducts=[
            ProductSummary("P001", "Brown Manga Lash", 50, 1250.00),
            ProductSummary("P002", "Natural Volume Lash", 35, 875.00)
        ]
    )

    # Write to Excel
    filename = writer.write_daily_data(sales_data)
    print(f"Excel file written: {filename}")

    # Read back
    data = writer.read_monthly_data("2026-06")
    print(f"Read {len(data)} records for 2026-06")