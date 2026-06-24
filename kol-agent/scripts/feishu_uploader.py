#!/usr/bin/env python3
"""
Feishu 多维表格上传模块
将视频数据上传至飞书多维表格
"""

import json
import os
import time
import urllib.request
from pathlib import Path

# 复用 feishu_refresh_token.py 的刷新逻辑
from feishu_refresh_token import refresh_user_access_token

CONFIG_PATH = Path(__file__).parent.parent / "tiktok-config.json"
CREATE_BITABLE_URL = "https://open.feishu.cn/open-apis/bitable/v1/apps"
BATCH_CREATE_URL = "https://open.feishu.cn/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/records/batch_create"
LIST_FIELDS_URL = "https://open.feishu.cn/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/fields"
DELETE_FIELD_URL = "https://open.feishu.cn/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/fields/{field_id}"
CREATE_FIELD_URL = "https://open.feishu.cn/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/fields"
LIST_RECORDS_URL = "https://open.feishu.cn/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/records"
BATCH_DELETE_RECORDS_URL = "https://open.feishu.cn/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/records/batch_delete"
BATCH_SIZE = 1000  # 每批最多1000条记录

# Excel 列名到飞书字段类型的映射
FIELD_TYPE_MAP = {
    "Video ID": {"type": 1, "field_name": "Video ID"},  # 文本
    "Title": {"type": 1, "field_name": "Title"},
    "Username": {"type": 1, "field_name": "Username"},
    "Creator User Name": {"type": 1, "field_name": "Creator User Name"},
    "Creator Nick Name": {"type": 1, "field_name": "Creator Nick Name"},
    "Creator Author Type": {"type": 1, "field_name": "Creator Author Type"},
    "Video Post Time": {"type": 1, "field_name": "Video Post Time"},
    "Duration (sec)": {"type": 2, "field_name": "Duration (sec)"},  # 数字
    "Hashtags": {"type": 1, "field_name": "Hashtags"},
    "GMV Amount": {"type": 2, "field_name": "GMV Amount"},
    "GMV Currency": {"type": 1, "field_name": "GMV Currency"},
    "GPM Amount": {"type": 2, "field_name": "GPM Amount"},
    "GPM Currency": {"type": 1, "field_name": "GPM Currency"},
    "Avg Customers": {"type": 2, "field_name": "Avg Customers"},
    "SKU Orders": {"type": 2, "field_name": "SKU Orders"},
    "Items Sold": {"type": 2, "field_name": "Items Sold"},
    "Views": {"type": 2, "field_name": "Views"},
    "Click Through Rate": {"type": 1, "field_name": "Click Through Rate"},
    "Products": {"type": 1, "field_name": "Products"},
}


class FeishuUploader:
    """飞书多维表格上传器"""

    def __init__(self, config_path: str = None):
        config_path = config_path or str(CONFIG_PATH)
        with open(config_path, "r") as f:
            self.config = json.load(f)

        feishu_config = self.config.get("feishu", {})
        self.app_id = feishu_config.get("app_id")
        self.app_secret = feishu_config.get("app_secret")
        self.user_access_token = feishu_config.get("user_access_token")
        self.token_expires_at = feishu_config.get("user_access_token_expires_at", 0)
        self.refresh_token = feishu_config.get("refresh_token")
        self.refresh_token_expires_at = feishu_config.get("refresh_token_expires_in", 0)

    def is_token_valid(self) -> bool:
        """检查 user_access_token 是否有效（提前 60 秒刷新）"""
        if not self.user_access_token:
            return False
        return time.time() < self.token_expires_at - 60

    def get_access_token(self) -> str:
        """获取有效的 user_access_token"""
        if not self.is_token_valid():
            self._refresh_token()
        return self.user_access_token

    def _refresh_token(self):
        """刷新 user_access_token（复用 feishu_refresh_token.py）"""
        result = refresh_user_access_token(self.app_id, self.app_secret, self.refresh_token)

        if result.get("code") == 0:
            self.user_access_token = result.get("access_token")
            expire = result.get("expires_in", 7200)
            self.token_expires_at = int(time.time()) + expire

            # 保存到配置
            self.config["feishu"]["user_access_token"] = self.user_access_token
            self.config["feishu"]["user_access_token_expires_at"] = self.token_expires_at

            if result.get("refresh_token"):
                self.refresh_token = result.get("refresh_token")
                self.config["feishu"]["refresh_token"] = self.refresh_token

            if result.get("refresh_token_expires_in"):
                self.refresh_token_expires_at = result.get("refresh_token_expires_in")
                self.config["feishu"]["refresh_token_expires_in"] = self.refresh_token_expires_at

            with open(CONFIG_PATH, "w") as f:
                json.dump(self.config, f, indent=2)
        else:
            raise Exception(f"Token 刷新失败: code={result.get('code')}, msg={result.get('error_description', result.get('msg'))}")

    def create_bitable(self, name: str) -> dict:
        """创建多维表格"""
        token = self.get_access_token()

        feishu_config = self.config.get("feishu", {})
        folder_token = feishu_config.get("folder_token")

        payload_dict = {"name": name}
        if folder_token:
            payload_dict["folder_token"] = folder_token

        payload = json.dumps(payload_dict).encode("utf-8")

        headers = {
            "Content-Type": "application/json; charset=utf-8",
            "Authorization": f"Bearer {token}"
        }

        req = urllib.request.Request(
            CREATE_BITABLE_URL,
            data=payload,
            headers=headers,
            method="POST"
        )

        try:
            with urllib.request.urlopen(req, timeout=30) as response:
                result = json.loads(response.read().decode())
        except urllib.error.HTTPError as e:
            error_body = e.read().decode("utf-8")
            raise Exception(f"创建多维表格失败: HTTP {e.code}, body={error_body}")

        if result.get("code") == 0:
            data = result.get("data", {}).get("app", {})
            return {
                "app_token": data.get("app_token"),
                "table_id": data.get("default_table_id"),
                "name": data.get("name"),
                "url": data.get("url")
            }
        else:
            raise Exception(f"创建多维表格失败: code={result.get('code')}, msg={result.get('msg')}")

    def list_fields(self, app_token: str, table_id: str) -> list:
        """列出表格所有字段"""
        token = self.get_access_token()

        url = LIST_FIELDS_URL.format(app_token=app_token, table_id=table_id)

        headers = {"Authorization": f"Bearer {token}"}

        req = urllib.request.Request(url, headers=headers, method="GET")

        with urllib.request.urlopen(req, timeout=30) as response:
            result = json.loads(response.read().decode())

        if result.get("code") == 0:
            return result.get("data", {}).get("items", [])
        else:
            raise Exception(f"列出字段失败: code={result.get('code')}, msg={result.get('msg')}")

    def delete_field(self, app_token: str, table_id: str, field_id: str):
        """删除字段"""
        token = self.get_access_token()

        url = DELETE_FIELD_URL.format(app_token=app_token, table_id=table_id, field_id=field_id)

        headers = {"Authorization": f"Bearer {token}"}

        req = urllib.request.Request(url, headers=headers, method="DELETE")

        with urllib.request.urlopen(req, timeout=30) as response:
            result = json.loads(response.read().decode())

        if result.get("code") != 0:
            raise Exception(f"删除字段失败: code={result.get('code')}, msg={result.get('msg')}")

    def create_field(self, app_token: str, table_id: str, field_name: str, field_type: int = 1):
        """创建字段 (type: 1=文本, 2=数字)"""
        token = self.get_access_token()

        payload = json.dumps({
            "field_name": field_name,
            "type": field_type
        }).encode("utf-8")

        headers = {
            "Content-Type": "application/json; charset=utf-8",
            "Authorization": f"Bearer {token}"
        }

        url = CREATE_FIELD_URL.format(app_token=app_token, table_id=table_id)
        req = urllib.request.Request(url, data=payload, headers=headers, method="POST")

        with urllib.request.urlopen(req, timeout=30) as response:
            result = json.loads(response.read().decode())

        if result.get("code") != 0:
            raise Exception(f"创建字段失败: code={result.get('code')}, msg={result.get('msg')}")

    def sync_fields(self, app_token: str, table_id: str, excel_headers: list):
        """同步表格字段：删除默认字段，创建新字段"""
        print(f"[Feishu] 同步字段...")

        # 1. 列出当前字段
        existing_fields = self.list_fields(app_token, table_id)
        print(f"[Feishu] 当前字段数: {len(existing_fields)}")

        # 2. 删除所有非主字段
        for field in existing_fields:
            field_name = field.get("field_name", "")
            is_primary = field.get("is_primary_field", False)
            # 主字段不能删除
            if is_primary:
                print(f"[Feishu]   保留主字段: {field_name}")
                continue
            if field_name and field_name != "_UUid_":
                print(f"[Feishu]   删除字段: {field_name} ({field.get('field_id')})")
                try:
                    self.delete_field(app_token, table_id, field.get("field_id"))
                except Exception as e:
                    if "Primary Field" in str(e):
                        print(f"[Feishu]   跳过主字段: {field_name}")
                    else:
                        raise

        # 3. 根据 Excel 表头创建新字段
        for header in excel_headers:
            field_info = FIELD_TYPE_MAP.get(header, {"type": 1})
            field_type = field_info.get("type", 1)
            print(f"[Feishu]   创建字段: {header} (type={field_type})")
            self.create_field(app_token, table_id, header, field_type)

        print(f"[Feishu] ✓ 字段同步完成，共 {len(excel_headers)} 个字段")

    def upload_batch(self, app_token: str, table_id: str, records: list) -> int:
        """批量上传记录"""
        token = self.get_access_token()

        url = BATCH_CREATE_URL.format(app_token=app_token, table_id=table_id)

        payload = json.dumps({"records": records}).encode("utf-8")

        headers = {
            "Content-Type": "application/json; charset=utf-8",
            "Authorization": f"Bearer {token}"
        }

        req = urllib.request.Request(url, data=payload, headers=headers, method="POST")

        try:
            with urllib.request.urlopen(req, timeout=60) as response:
                result = json.loads(response.read().decode())
        except urllib.error.HTTPError as e:
            error_body = e.read().decode("utf-8")
            raise Exception(f"批量上传失败: HTTP {e.code}, body={error_body}")

        if result.get("code") == 0:
            return len(result.get("data", {}).get("records", []))
        else:
            raise Exception(f"批量上传失败: code={result.get('code')}, msg={result.get('msg')}")

    def list_records(self, app_token: str, table_id: str, page_size: int = 100) -> list:
        """获取表格所有记录（用于删除空白记录）"""
        token = self.get_access_token()
        all_records = []
        page_token = None

        while True:
            url = LIST_RECORDS_URL.format(app_token=app_token, table_id=table_id)
            params = f"?page_size={page_size}"
            if page_token:
                params += f"&page_token={page_token}"

            headers = {"Authorization": f"Bearer {token}"}
            req = urllib.request.Request(url + params, headers=headers, method="GET")

            with urllib.request.urlopen(req, timeout=30) as response:
                result = json.loads(response.read().decode())

            if result.get("code") != 0:
                raise Exception(f"获取记录失败: code={result.get('code')}, msg={result.get('msg')}")

            items = result.get("data", {}).get("items", [])
            all_records.extend(items)

            page_token = result.get("data", {}).get("page_token")
            if not page_token:
                break

        return all_records

    def delete_records(self, app_token: str, table_id: str, record_ids: list) -> int:
        """批量删除记录"""
        token = self.get_access_token()
        url = BATCH_DELETE_RECORDS_URL.format(app_token=app_token, table_id=table_id)

        payload = json.dumps({"records": record_ids}).encode("utf-8")
        headers = {
            "Content-Type": "application/json; charset=utf-8",
            "Authorization": f"Bearer {token}"
        }

        req = urllib.request.Request(url, data=payload, headers=headers, method="POST")

        try:
            with urllib.request.urlopen(req, timeout=30) as response:
                result = json.loads(response.read().decode())
        except urllib.error.HTTPError as e:
            error_body = e.read().decode("utf-8")
            raise Exception(f"批量删除失败: HTTP {e.code}, body={error_body}")

        if result.get("code") == 0:
            deleted_list = result.get("data", {}).get("records", [])
            return len([r for r in deleted_list if r.get("deleted")])
        else:
            raise Exception(f"批量删除失败: code={result.get('code')}, msg={result.get('msg')}")

    def delete_empty_records(self, app_token: str, table_id: str) -> int:
        """删除表格中的空白记录（新建表格会自带空白记录）"""
        print(f"[Feishu] 检查空白记录...")

        records = self.list_records(app_token, table_id)
        if not records:
            print(f"[Feishu]   无记录")
            return 0

        # 检查哪些记录是空白的（所有字段都为空）
        empty_record_ids = []
        for record in records:
            fields = record.get("fields", {})
            # 如果所有字段都为空，则是空白记录
            if all(v is None or v == "" for v in fields.values()):
                empty_record_ids.append(record.get("record_id"))

        if not empty_record_ids:
            print(f"[Feishu]   无空白记录")
            return 0

        print(f"[Feishu]   发现 {len(empty_record_ids)} 条空白记录，开始删除...")

        # 分批删除（每批最多500条）
        total_deleted = 0
        for i in range(0, len(empty_record_ids), 500):
            batch = empty_record_ids[i:i + 500]
            deleted = self.delete_records(app_token, table_id, batch)
            total_deleted += deleted
            print(f"[Feishu]     已删除 {total_deleted}/{len(empty_record_ids)} 条")

        print(f"[Feishu]   ✓ 空白记录删除完成，共 {total_deleted} 条")
        return total_deleted

    def upload_excel_to_bitable(self, file_path: str, date: str) -> dict:
        """将 Excel 文件上传到飞书多维表格"""
        import openpyxl

        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")

        print(f"[Feishu] 开始上传文件: {file_path}")
        print(f"[Feishu] 数据日期: {date}")

        # 读取 Excel
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # 获取表头和数据
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)

        # 转换为记录格式
        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            fields = {}
            for i, value in enumerate(row):
                if i < len(headers) and headers[i]:
                    field_name = str(headers[i])
                    # 转换数值类型
                    if field_info := FIELD_TYPE_MAP.get(headers[i]):
                        if field_info.get("type") == 2 and value is not None:
                            # 数字字段：尝试转换为数值
                            try:
                                value = float(str(value).replace(",", ""))
                            except (ValueError, TypeError):
                                pass
                    fields[field_name] = value
            if fields:
                records.append({"fields": fields})

        total_count = len(records)
        print(f"[Feishu] Excel 总记录数: {total_count}")

        # 创建多维表格
        bitable_name = f"视频数据_{date}"
        print(f"[Feishu] 创建多维表格: {bitable_name}")
        bitable_info = self.create_bitable(bitable_name)
        app_token = bitable_info["app_token"]
        table_id = bitable_info["table_id"]
        bitable_url = bitable_info["url"]

        print(f"[Feishu] ✓ 多维表格创建成功")
        print(f"[Feishu]   URL: {bitable_url}")

        # 同步字段
        self.sync_fields(app_token, table_id, headers)

        # 删除空白记录（新建表格会自带10条空白记录）
        self.delete_empty_records(app_token, table_id)

        # 分批上传
        uploaded_count = 0
        batch_num = 0

        for i in range(0, total_count, BATCH_SIZE):
            batch = records[i:i + BATCH_SIZE]
            batch_num += 1
            print(f"[Feishu] 上传第 {batch_num} 批 ({len(batch)} 条)...")

            try:
                count = self.upload_batch(app_token, table_id, batch)
                uploaded_count += count
                print(f"[Feishu]   已上传: {uploaded_count}/{total_count}")
            except Exception as e:
                print(f"[Feishu]   上传失败: {e}")
                raise

        # 校验
        if uploaded_count != total_count:
            print(f"[Feishu] ⚠️ 警告: 上传记录数 ({uploaded_count}) 与 Excel 行数 ({total_count}) 不一致")

        print(f"[Feishu] ✓ 上传完成")
        print(f"[Feishu]   Bitable URL: {bitable_url}")
        print(f"[Feishu]   记录数: {uploaded_count}/{total_count}")

        return {
            "success": True,
            "bitable_url": bitable_url,
            "record_count": uploaded_count,
            "total_count": total_count,
            "date": date
        }


def upload_to_feishu(file_path: str, date: str = None) -> dict:
    """快捷函数：上传 Excel 到飞书多维表格"""
    uploader = FeishuUploader()

    if not date:
        import re
        match = re.search(r"(\d{4}-\d{2}-\d{2})", file_path)
        date = match.group(1) if match else "unknown"

    return uploader.upload_excel_to_bitable(file_path, date)


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("用法: python3 feishu_uploader.py <file_path> [date]")
        sys.exit(1)

    file_path = sys.argv[1]
    date = sys.argv[2] if len(sys.argv) > 2 else None

    result = upload_to_feishu(file_path, date)
    sys.exit(0 if result.get("success") else 1)
